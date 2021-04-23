# -*- coding: utf-8 -*-
from numpy import genfromtxt, array
from utils.os_control import *
import openpyxl


def read_data(graph, root_dir):
    try:
        data = genfromtxt(find_file(root_dir + graph),
                          dtype=int, delimiter='\t', comments='%')
        data = data[data[:, 3].argsort()]
    except:
        data = genfromtxt(find_file(root_dir + graph),
                          dtype=int, delimiter=' ', comments='%')
        data = data[data[:, 3].argsort()]
    return data


def cut_graph(graph, STEP_NUM, min_len, max_len, root_dir, target_dir):
    print('-'*58)
    print('reading data {}...'.format(graph[:-1]), end='')
    data = read_data(graph, root_dir)
    print('finished')
    s_nodes = data[:, 0]
    t_nodes = data[:, 1]
    timestamps = data[:, 3]
    time_min, time_max = min(timestamps), max(timestamps)
    time_len = time_max - time_min
    time_step = int(time_len * STEP_NUM)
    data_list_groups = []
    data_list = []
    tmp_time_max = time_step + time_min
    data_len = len(timestamps)
    print('classification datas...', end='')
    for tmp_idx in range(data_len):
        tmp_list = [s_nodes[tmp_idx], t_nodes[tmp_idx]]
        if(timestamps[tmp_idx] <= tmp_time_max):
            if(tmp_list not in data_list):
                data_list.append(tmp_list)
        else:
            tmp_time_max += time_step
            if(min_len <= len(data_list) <= max_len):
                data_list_groups.append(data_list)
            data_list = []
            if(timestamps[tmp_idx] <= tmp_time_max):
                if(tmp_list not in data_list):
                    data_list.append(tmp_list)
    if(min_len <= len(data_list) <= max_len):
        data_list_groups.append(data_list)
    target_graph_dir = target_dir + graph
    rmdir(target_graph_dir)
    ensure_dir(target_graph_dir)

    data_list_groups_final = []
    for graph in data_list_groups:
        tmp_edge_list = []
        edge_set_list = []
        for edge in graph:
            s_node, t_node = edge[0], edge[1]
            out_of_graph = True
            for edge_set_idx in range(len(edge_set_list)):
                if(s_node in edge_set_list[edge_set_idx] or t_node in edge_set_list[edge_set_idx]):
                    edge_set_list[edge_set_idx].update(edge)
                    tmp_edge_list[edge_set_idx].append(edge)
                    out_of_graph = False
            if(out_of_graph):
                edge_set_list.append(set(edge))
                tmp_edge_list.append([edge])

        while(1):
            len_of_list = len(edge_set_list)
            continue_flag = False
            for set1_idx in range(len_of_list):
                for set2_idx in range(set1_idx+1, len_of_list):
                    if(edge_set_list[set1_idx] & edge_set_list[set2_idx]):
                        edge_set_list[set1_idx] = edge_set_list[set1_idx] | edge_set_list[set2_idx]
                        tmp_edge_list[set1_idx].extend(tmp_edge_list[set2_idx])
                        del edge_set_list[set2_idx]
                        del tmp_edge_list[set2_idx]
                        continue_flag = True
                        break
                if(continue_flag):
                    break
            if(continue_flag):
                continue
            else:
                break
        for i in tmp_edge_list:
            if(min_len < len(i) < max_len):
                data_list_groups_final.append(i)
    print('finished')
    print('saving datas...', end='')
    data_list_group_len = len(data_list_groups_final)
    for idx_group in range(data_list_group_len):
        with open(target_graph_dir + '{}.txt'.format(idx_group), 'w') as f:
            for nodes in data_list_groups_final[idx_group]:
                f.write(str(nodes[0]) + '\t' + str(nodes[1]) + '\n')
    print('finished')
    return data_list_group_len


def data_cut(graph_box, gap, min_num, max_num, root_dir, target_dir):
    rmdir(target_dir)
    ensure_dir(target_dir)
    graph_num_box = []
    for graph in graph_box:
        graph_num_box.append(
            cut_graph(graph + '/', gap, min_num, max_num, root_dir, target_dir))
    return graph_num_box


def cal_graph(graph_name, root_dir):
    graph = graph_name + '/'
    data = array([])
    try:
        data = genfromtxt(find_file(root_dir + graph),
                          dtype=int, delimiter='\t', comments='%')
        s_nodes = data[:, 0]
        t_nodes = data[:, 1]
    except:
        data = genfromtxt(find_file(root_dir + graph),
                          dtype=int, delimiter=' ', comments='%')
        s_nodes = data[:, 0]
        t_nodes = data[:, 1]
    return len(set(s_nodes)), len(set(t_nodes)), len(data)


def data_cal(graph_box, gap, min_num, max_num, timestamp):
    root_dir = './output/datas_origin/'
    ensure_dir('./output/output/')
    ensure_dir('./output/output/{}/'.format(timestamp))
    target_dir = './output/output/{}/datas/'.format(timestamp)
    excel_path = './output/output/{}/dataset_info.xlsx'.format(timestamp)
    graph_num_box = data_cut(graph_box, gap, min_num,
                             max_num, root_dir, target_dir)
    wb = openpyxl.Workbook()
    ws = wb.active
    idx = 1
    ws['A{}'.format(idx)] = 'Data'
    ws['B{}'.format(idx)] = 'S Vertex'
    ws['C{}'.format(idx)] = 'T Vertex'
    ws['D{}'.format(idx)] = 'Edge'
    ws['E{}'.format(idx)] = '#Graph'
    idx += 1
    print('\n', '='*58)
    print('||{:^54}||'.format('Generated dataset statistics'))
    print('||' + '='*54 + '||')
    print('||{:<30}|{:<5}|{:<5}|{:<6}|{:<4}||'.format(
        'Dataset', '#.S', '#,T', '#.E', '#.G'))
    for graph, num in zip(graph_box, graph_num_box):
        s_node_len, t_node_len, data_len = cal_graph(graph, root_dir)
        print('||' + '-'*54 + '||')
        print('||{:<30}|{:<5}|{:<5}|{:<6}|{:<4}||'.format(
            graph, s_node_len, t_node_len, data_len, num))
        ws['A{}'.format(idx)] = graph
        ws['B{}'.format(idx)] = s_node_len
        ws['C{}'.format(idx)] = t_node_len
        ws['D{}'.format(idx)] = data_len
        ws['E{}'.format(idx)] = num
        idx += 1
    print('='*58, '\n')
    while(1):
        try:
            # python can't save excel file while user has been opened it.
            wb.save(excel_path)
            break
        except:
            input('please close the excel and try again!')
